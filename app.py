import streamlit as st
import pandas as pd
import re
import io
import os
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

# ==========================================
# CONFIGURAÇÃO GERAL
# ==========================================
st.set_page_config(
    page_title="Conciliação Contábil",
    layout="wide",
    page_icon="📊",
    initial_sidebar_state="collapsed"
)

st.markdown("""
    <style>
        .block-container {padding-top: 2rem; padding-bottom: 2rem;}
        div[data-testid="stFileUploader"] section {padding: 10px;}
        h1 {font-size: 1.8rem;}
        h3 {font-size: 1.2rem;}
    </style>
""", unsafe_allow_html=True)

# ==========================================
# 1. FUNÇÕES DE LIMPEZA E FORMATAÇÃO
# ==========================================
def gerar_chave_padronizada(texto_conta):
    if not isinstance(texto_conta, str): return None
    texto_conta = texto_conta.strip()
    
    if '.' in texto_conta and len(texto_conta) > 12:
        partes = texto_conta.split('.')
        maior_parte = ""
        for p in partes:
            limpo = re.sub(r'\D', '', p)
            if len(limpo) > len(maior_parte): maior_parte = limpo
        if len(maior_parte) > 4: texto_conta = maior_parte
    elif '/' in texto_conta:
        texto_conta = texto_conta.split('/')[-1]
            
    parte_numerica = re.sub(r'\D', '', texto_conta)
    if not parte_numerica: return None
    
    return parte_numerica[-7:].zfill(7)

def limpar_valor_monetario(valor_str):
    if not isinstance(valor_str, str): return 0.0
    valor_upper = valor_str.upper()
    eh_negativo = 'D' in valor_upper or 'DEB' in valor_upper or '-' in valor_str or '(' in valor_str
    
    limpo = re.sub(r'[^\d,\.]', '', valor_str)
    
    try:
        if not limpo: return 0.0
        if ',' in limpo and '.' in limpo:
             limpo = limpo.replace('.', '').replace(',', '.')
        elif ',' in limpo:
             limpo = limpo.replace(',', '.')
        
        valor_float = float(limpo)
        return -valor_float if eh_negativo else valor_float
    except ValueError:
        return 0.0

def formatar_moeda_br(valor):
    if pd.isna(valor): return "0,00"
    return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def extrair_valor_monetario_flex(texto):
    """
    Função blindada para o layout da Caixa Econômica e Banco do Brasil.
    Evita misturar cotas com moeda e lida perfeitamente com 3.952,38C ou 123,45D.
    """
    match = re.search(r"(-?\d{1,3}(?:\.\d{3})*,\d{2})(?!\d)\s*([CDcd\-])?", texto)
    if not match: return 0.0
    
    numero_str = match.group(1)
    sulfixo = match.group(2)
    
    limpo = numero_str.replace('.', '').replace(',', '.')
    try:
        valor = float(limpo)
    except:
        return 0.0
        
    texto_upper = texto.upper()
    is_negative = False
    
    if sulfixo and sulfixo.upper() in ['D', '-']:
        is_negative = True
    elif "-" in numero_str:
        is_negative = True
    elif (re.search(r"\b(D|DEB|DEBITO)\b", texto_upper) and "DIAS" not in texto_upper):
        is_negative = True
        
    return -abs(valor) if is_negative else abs(valor)

# ==========================================
# 2. MOTOR DE LEITURA DE PDF
# ==========================================
def extrair_pdf_melhorado(arquivo, tipo_extrato):
    try:
        doc = fitz.open(stream=arquivo.read(), filetype="pdf")
        texto_completo = ""
        for pag in doc:
            texto_completo += pag.get_text() + "\n"
        doc.close()
        
        linhas = texto_completo.split('\n')
        
        conta_encontrada = "N/A"
        padroes_conta = [
            r"Conta:\s*(\d{4}\/\d{3,4}\/[\d\-]+)", 
            r"Conta\s*Vinculada:\s*(\d{4}\/\d{3,4}\/[\d\-]+)", 
            r"Conta\s*Corrente\s*[:\s]*([\d\.\-\/]+)",   
            r"Conta\s*[:\s]*([\d\.\-\/]+)",                
            r"Agência.*?Conta.*?([\d\.\-]{5,})",            
            r"C\/C\s*[:\s]*([\d\.\-\/]+)"                 
        ]
        
        for p in padroes_conta:
            match = re.search(p, texto_completo, re.IGNORECASE)
            if match:
                conta_raw = match.group(1).strip()
                if len(re.sub(r'\D', '', conta_raw)) > 4:
                    conta_encontrada = conta_raw
                    break
        
        if conta_encontrada == "N/A":
            cabecalho = "\n".join(linhas[:25]) 
            match_solto = re.search(r"(\d{4,6}-\d)", cabecalho)
            if match_solto: conta_encontrada = match_solto.group(1)

        saldo_final = 0.0
        rendimento_total = 0.0

        leu_rendimento_deste_fundo = False

        for i, linha in enumerate(linhas):
            linha_upper = linha.upper().strip()
            
            if "SALDO ANTERIOR" in linha_upper:
                leu_rendimento_deste_fundo = False
            
            # --- 1. CAPTURA DE SALDO FINAL ---
            gatilhos_saldo = ["SALDO FINAL", "SALDO TOTAL", "SALDO ATUAL", "SALDO EM", "SALDO LÍQUIDO", "SALDO BRUTO", "VALOR LIQUIDO", "TOTAL DISPONIVEL", "POSICAO EM", "TOTAL EM COTAS", "S A L D O"]
            ignorar_saldo = ["ANTERIOR", "BLOQUEADO", "PROVISORIO", "RENDIMENTO", "RENTABILIDADE", "="]
            
            if any(g in linha_upper for g in gatilhos_saldo) and not any(ign in linha_upper for ign in ignorar_saldo):
                v = 0.0
                if re.search(r"\d", linha_upper):
                    v = extrair_valor_monetario_flex(linha_upper)
                
                if v == 0.0:
                    for j in range(1, 5):
                        if i + j < len(linhas):
                            v_temp = extrair_valor_monetario_flex(linhas[i+j])
                            if v_temp != 0.0:
                                v = v_temp
                                break
                
                if v != 0.0: 
                    saldo_final += v  
                    
            # --- 2. CAPTURA DE RENDIMENTO ---
            if tipo_extrato == 'INV' and not leu_rendimento_deste_fundo:
                gatilhos_rend = ["RENDIMENTO BRUTO", "RENTABILIDADE", "RENDIMENTO NO MÊS", "RENDIMENTO LIQUIDO", "RENTAB."]
                
                if any(g in linha_upper for g in gatilhos_rend) and "ACUMULADO" not in linha_upper and "ANO" not in linha_upper:
                    v = 0.0
                    if re.search(r"\d", linha_upper):
                        v = extrair_valor_monetario_flex(linha_upper)
                        
                    if v == 0.0:
                        for j in range(1, 5):
                            if i + j < len(linhas):
                                v_temp = extrair_valor_monetario_flex(linhas[i+j])
                                if v_temp != 0.0:
                                    v = v_temp
                                    break
                    
                    if v != 0.0 and abs(v) < 50000000:
                         rendimento_total += v 
                         leu_rendimento_deste_fundo = True

        if saldo_final == 0.0 and ("NAO HOUVE MOVIMENTO" in texto_completo.upper() or "SEM MOVIMENTO" in texto_completo.upper()):
             match_ant = re.search(r"(?:SALDO ANTERIOR|SALDO).*?(\d{1,3}(?:\.\d{3})*,\d{2})", texto_completo, re.IGNORECASE | re.DOTALL)
             if match_ant: saldo_final = extrair_valor_monetario_flex(match_ant.group(0))

        if saldo_final == 0.0 and tipo_extrato == 'INV':
            match_last = re.findall(r"(?:TOTAL|SALDO|ATUAL|LÍQUIDO).*?(-?\d{1,3}(?:\.\d{3})*,\d{2})(?!\d)", texto_completo, re.IGNORECASE | re.DOTALL)
            if match_last: saldo_final = extrair_valor_monetario_flex(match_last[-1])

        texto_limpo = texto_completo[:300].replace('\n', ' ').replace(';', ',')
        return {"Conta": conta_encontrada, "Saldo": saldo_final, "Rendimento": rendimento_total, "Texto_Raw": texto_limpo}
    except Exception as e:
        return {"Conta": "Erro", "Saldo": 0.0, "Rendimento": 0.0, "Texto_Raw": str(e)}

def carregar_depara():
    import streamlit as st
    
    diretorio_atual = os.path.dirname(os.path.abspath(__file__))
    caminho_arquivo = os.path.join(diretorio_atual, "depara", "DEPARA_CONTAS BANCÁRIAS_CEF.xlsx")
    
    try:
        df_depara = pd.read_excel(caminho_arquivo, sheet_name="2025_JUNHO (2)", dtype=str, engine='openpyxl')
        
        if len(df_depara.columns) != 2:
            df_depara = df_depara.iloc[:, :2]
            
        df_depara.columns = ['Conta Antiga', 'Conta Nova']
        
        df_depara['Chave Antiga'] = df_depara['Conta Antiga'].apply(gerar_chave_padronizada)
        df_depara['Chave Nova'] = df_depara['Conta Nova'].apply(gerar_chave_padronizada)
        
        # BLINDAGEM: Remove linhas em branco que podem bugar a criação do dicionário
        df_depara = df_depara.dropna(subset=['Chave Antiga', 'Chave Nova'])
        
        return df_depara
        
    except FileNotFoundError:
        st.error(f"⚠️ O arquivo De-Para não foi encontrado neste caminho:\n{caminho_arquivo}")
        return pd.DataFrame()
    except Exception as e:
        st.error(f"⚠️ Erro ao tentar ler a planilha De-Para:\n{e}")
        return pd.DataFrame()

# ==========================================
# 3. LEITURA CONTÁBIL
# ==========================================
def processar_contabil(arquivo, tipo='SALDO'):
    if arquivo is None: return pd.DataFrame()
    try:
        df = pd.read_csv(arquivo, encoding='latin-1', sep=';', header=1, dtype=str)
        col_chave = None
        possiveis_chaves = ['Domicílio bancário', 'Conta', 'Nº Conta', 'Descrição', 'Conta Contabil']
        for col in df.columns:
            for p in possiveis_chaves:
                if p.lower() in str(col).lower(): col_chave = col; break
        
        if not col_chave:
            arquivo.seek(0)
            df = pd.read_csv(arquivo, encoding='latin-1', sep=';', header=0, dtype=str)
            for col in df.columns:
                for p in possiveis_chaves:
                    if p.lower() in str(col).lower(): col_chave = col; break
        if not col_chave: return pd.DataFrame()

        col_valor = None
        possiveis_valores = ['Saldo Final', 'Saldo Atual', 'Movimento', 'Valor']
        for col in df.columns:
            if 'anterior' in str(col).lower(): continue
            for p in possiveis_valores:
                if p.lower() in str(col).lower(): col_valor = col; break
        if not col_valor: return pd.DataFrame()

        df['Chave Primaria'] = df[col_chave].apply(gerar_chave_padronizada)
        df = df.dropna(subset=['Chave Primaria'])
        df['Valor_Numerico'] = df[col_valor].astype(str).apply(limpar_valor_monetario)
        
        col_desc_original = col_chave 
        for col in df.columns:
            if ('descri' in str(col).lower() or 'nome' in str(col).lower()) and col != col_chave:
                col_desc_original = col
                break

        if tipo == 'SALDO':
            if any('contábil' in str(c).lower() for c in df.columns):
                col_contabil = next(c for c in df.columns if 'contábil' in str(c).lower())
                df_pivot = df.pivot_table(index='Chave Primaria', columns=col_contabil, values='Valor_Numerico', aggfunc='sum').reset_index()
                
                col_mov = next((c for c in df_pivot.columns if '1111119' in str(c) or 'Conta Movimento' in str(c) or 'MOVIMENTO' in str(c).upper()), None)
                col_app = next((c for c in df_pivot.columns if '1111150' in str(c) or 'Aplicação' in str(c) or 'APLICACAO' in str(c).upper()), None)
                
                df_res = pd.DataFrame()
                df_res['Chave Primaria'] = df_pivot['Chave Primaria']
                df_res['Saldo_Contabil_CC'] = df_pivot[col_mov].fillna(0) if col_mov else 0.0
                df_res['Saldo_Contabil_Aplic'] = df_pivot[col_app].fillna(0) if col_app else 0.0
                
                desc = df[['Chave Primaria', col_desc_original]].drop_duplicates(subset='Chave Primaria')
                df_res = df_res.merge(desc, on='Chave Primaria', how='left')
                df_res.rename(columns={col_desc_original: 'Descrição_ERP'}, inplace=True)
                return df_res
            else:
                df_agrup = df.groupby('Chave Primaria')['Valor_Numerico'].sum().reset_index()
                df_agrup.rename(columns={'Valor_Numerico': 'Saldo_Contabil_CC'}, inplace=True)
                df_agrup['Saldo_Contabil_Aplic'] = 0.0
                
                desc = df[['Chave Primaria', col_desc_original]].drop_duplicates(subset='Chave Primaria')
                df_agrup = df_agrup.merge(desc, on='Chave Primaria', how='left')
                df_agrup.rename(columns={col_desc_original: 'Descrição_ERP'}, inplace=True)
                return df_agrup

        elif tipo == 'RENDIMENTO':
            df_agrup = df.groupby('Chave Primaria')['Valor_Numerico'].sum().reset_index()
            df_agrup.rename(columns={'Valor_Numerico': 'Rendimento_Contabil'}, inplace=True)
            return df_agrup
    except Exception as e:
        return pd.DataFrame()

# ==========================================
# 4. CONSOLIDAÇÃO, DE-PARA E CLASSIFICAÇÃO
# ==========================================
def identificar_banco_por_texto(row):
    if pd.notna(row.get('Nome_Banco')) and str(row.get('Nome_Banco')) not in ['0', '0.0', 'nan', 'None']:
        return str(row['Nome_Banco']).upper()
    
    desc = str(row.get('Descrição_ERP', '')).upper()
    
    if 'BRASIL' in desc or 'BB ' in desc or 'BCO DO BRASIL' in desc:
        return "BANCO DO BRASIL"
    elif 'CAIXA' in desc or 'CEF' in desc or 'FEDERAL' in desc or 'ECONÔMICA' in desc:
        return "CAIXA ECONÔMICA"
    
    if '001' in desc: return "BANCO DO BRASIL"
    if '104' in desc: return "CAIXA ECONÔMICA"
    
    return desc

def executar_processo(file_saldos, file_rendim, lista_arquivos_bancarios):
    df_saldos = processar_contabil(file_saldos, 'SALDO')
    df_rendim = processar_contabil(file_rendim, 'RENDIMENTO')
    
    if df_saldos.empty:
        st.error("Erro na leitura do CSV de Saldos.")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    df_depara = carregar_depara()
    
    if not df_depara.empty:
        dicionario_depara = dict(zip(df_depara['Chave Antiga'], df_depara['Chave Nova']))
        
        # --- APLICA O DE-PARA USANDO MAP (Blindagem contra falha silenciosa do Pandas) ---
        df_saldos['Chave Primaria'] = df_saldos['Chave Primaria'].map(lambda x: dicionario_depara.get(x, x))
        df_saldos = df_saldos.groupby('Chave Primaria', as_index=False).agg({
            'Saldo_Contabil_CC': 'sum',
            'Saldo_Contabil_Aplic': 'sum',
            'Descrição_ERP': 'first'
        })
        
        if not df_rendim.empty:
            df_rendim['Chave Primaria'] = df_rendim['Chave Primaria'].map(lambda x: dicionario_depara.get(x, x))
            df_rendim = df_rendim.groupby('Chave Primaria', as_index=False).agg({
                'Rendimento_Contabil': 'sum'
            })

    df_contabil = df_saldos
    if not df_rendim.empty:
        df_contabil = pd.merge(df_saldos, df_rendim, on='Chave Primaria', how='outer').fillna(0)
    else:
        df_contabil['Rendimento_Contabil'] = 0.0

    dados_banco = []
    log_leitura = []

    for item in lista_arquivos_bancarios:
        f = item['arquivo']
        banco_nome = item['banco']
        tipo_extrato = item['tipo']
        
        res = extrair_pdf_melhorado(f, tipo_extrato)
        chave = gerar_chave_padronizada(res['Conta'])
        
        log_leitura.append({
            'Arquivo': f.name, 
            'Banco': banco_nome,
            'Conta Lida': res['Conta'], 
            'Chave Gerada': str(chave), 
            'Saldo': res['Saldo'], 
            'Rendimento': res['Rendimento'] if tipo_extrato == 'INV' else 0.0
        })

        if chave: 
            dados_banco.append({
                'Chave Primaria': chave, 
                'Nome_Banco': banco_nome,
                'Saldo_Banco_CC': res['Saldo'] if tipo_extrato == 'CC' else 0.0,
                'Saldo_Banco_Aplic': res['Saldo'] if tipo_extrato == 'INV' else 0.0, 
                'Rendimento_Banco': res['Rendimento'] if tipo_extrato == 'INV' else 0.0
            })

    df_log = pd.DataFrame(log_leitura)
    
    if dados_banco:
        df_banco = pd.DataFrame(dados_banco)
        
        # Blindagem extra no banco com map
        if not df_depara.empty:
             df_banco['Chave Primaria'] = df_banco['Chave Primaria'].map(lambda x: dicionario_depara.get(x, x))
             
        df_banco = df_banco.groupby('Chave Primaria').agg({
            'Saldo_Banco_CC': 'sum',
            'Saldo_Banco_Aplic': 'sum',
            'Rendimento_Banco': 'sum',
            'Nome_Banco': 'first'
        }).reset_index()
    else:
        df_banco = pd.DataFrame(columns=['Chave Primaria', 'Saldo_Banco_CC', 'Saldo_Banco_Aplic', 'Rendimento_Banco', 'Nome_Banco'])

    df_final = pd.merge(df_contabil, df_banco, on='Chave Primaria', how='outer').fillna(0)

    df_final['Descrição'] = df_final.apply(identificar_banco_por_texto, axis=1)
    df_final['Descrição'] = df_final['Descrição'].astype(str).str.upper().replace(['NAN', 'NONE', '0', ''], '-')

    df_final['Diferenca_Saldo_CC'] = df_final['Saldo_Contabil_CC'] - df_final['Saldo_Banco_CC']
    df_final['Diferenca_Saldo_Aplic'] = df_final['Saldo_Contabil_Aplic'] - df_final['Saldo_Banco_Aplic']
    df_final['Diferenca_Rendimento'] = df_final['Rendimento_Contabil'] - df_final['Rendimento_Banco']

    cols = ['Descrição', 'Chave Primaria', 'Saldo_Contabil_CC', 'Saldo_Banco_CC', 'Diferenca_Saldo_CC',
            'Saldo_Contabil_Aplic', 'Saldo_Banco_Aplic', 'Diferenca_Saldo_Aplic',
            'Rendimento_Contabil', 'Rendimento_Banco', 'Diferenca_Rendimento']
    colunas_finais = [c for c in cols if c in df_final.columns]
    
    return df_final[colunas_finais], df_log, df_depara 

# ==========================================
# 5. GERADORES DE ARQUIVO (Excel e PDF)
# ==========================================

def to_excel_styled(df):
    output = io.BytesIO()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Conciliação"

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(bold=True, name="Calibri", size=11)
    border_style = Side(border_style="thin", color="000000")
    thin_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    alignment_center = Alignment(horizontal="center", vertical="center")
    number_fmt = '#,##0.00'
    
    headers_lvl0 = [c[0] for c in df.columns]
    headers_lvl1 = [c[1] for c in df.columns]
    
    for col_idx, header in enumerate(headers_lvl0, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment_center
        cell.border = thin_border
    
    for col_idx, header in enumerate(headers_lvl1, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment_center
        cell.border = thin_border

    rows = dataframe_to_rows(df, index=False, header=False)
    for r_idx, row in enumerate(rows, 3):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            
            if isinstance(value, (int, float)):
                cell.number_format = number_fmt
                if value < -0.01:
                    cell.font = Font(color="FF0000")
                
                header_name = headers_lvl1[c_idx-1]
                if "Diferença" in header_name and abs(value) > 0.01:
                     cell.font = Font(color="FF0000", bold=True)
            
    for i in range(1, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    wb.save(output)
    return output.getvalue()

def to_pdf(df):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    title_style.alignment = 1 
    elements.append(Paragraph("Relatório de Conciliação Contábil", title_style))
    elements.append(Spacer(1, 12))

    headers = [f"{c[0]}\n{c[1]}" for c in df.columns]
    
    data = [headers]
    
    for index, row in df.iterrows():
        row_list = []
        for col_name, val in row.items():
            if isinstance(val, (int, float)):
                row_list.append(formatar_moeda_br(val))
            else:
                row_list.append(str(val))
        data.append(row_list)

    col_widths = [100, 60] + [70] * 9 
    t = Table(data, colWidths=None)

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 1), (1, -1), 'LEFT'), 
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'), 
    ])
    
    for i in range(1, len(data)):
        if i % 2 == 0:
            bg_color = colors.whitesmoke
        else:
            bg_color = colors.white
        style.add('BACKGROUND', (0, i), (-1, i), bg_color)
        
        for j, val in enumerate(data[i]):
            if j > 1 and ("-" in val or "(" in val) and val != "0,00":
                 style.add('TEXTCOLOR', (j, i), (j, i), colors.red)

    t.setStyle(style)
    elements.append(t)
    
    doc.build(elements)
    return buffer.getvalue()

# ==========================================
# 6. INTERFACE DO USUÁRIO
# ==========================================
st.title("Sistema de Conciliação Financeira")
st.markdown("---")

col_left, col_right = st.columns(2)

with col_left:
    with st.container(border=True):
        st.subheader("1. Arquivos Contábeis (CSV)")
        f_saldos = st.file_uploader("Conta corrente (.CSV) - Relatório Flexvision 013083", type='csv')
        f_rendim = st.file_uploader("Rendimentos (.CSV) - Relatório Flexvision 014387", type='csv')

with col_right:
    with st.container(border=True):
        st.subheader("2. Extratos Bancários (PDF)")
        f_bb_cc = st.file_uploader("🔵 Banco do Brasil - Conta Corrente", type='pdf', accept_multiple_files=True)
        f_bb_inv = st.file_uploader("🔵 Banco do Brasil - Investimentos", type='pdf', accept_multiple_files=True)
        
        st.divider()
        
        f_caixa_cc = st.file_uploader("🟠 Caixa Econômica - Conta Corrente", type='pdf', accept_multiple_files=True)
        f_caixa_inv = st.file_uploader("🟠 Caixa Econômica - Investimentos", type='pdf', accept_multiple_files=True)

st.markdown("<br>", unsafe_allow_html=True)
btn_processar = st.button("Processar Conciliação", type="primary", use_container_width=True)

if btn_processar:
    if not f_saldos:
        st.warning("⚠️ Obrigatório carregar o arquivo de Saldos (CSV).")
    else:
        lista_arquivos = []
        if f_bb_cc:
            for f in f_bb_cc: lista_arquivos.append({'arquivo': f, 'banco': 'BANCO DO BRASIL', 'tipo': 'CC'})
        if f_bb_inv:
            for f in f_bb_inv: lista_arquivos.append({'arquivo': f, 'banco': 'BANCO DO BRASIL', 'tipo': 'INV'})
        if f_caixa_cc:
            for f in f_caixa_cc: lista_arquivos.append({'arquivo': f, 'banco': 'CAIXA ECONÔMICA', 'tipo': 'CC'})
        if f_caixa_inv:
            for f in f_caixa_inv: lista_arquivos.append({'arquivo': f, 'banco': 'CAIXA ECONÔMICA', 'tipo': 'INV'})
        
        with st.spinner("Lendo arquivos e cruzando dados..."):
            df_final, df_log, df_depara = executar_processo(f_saldos, f_rendim, lista_arquivos)
            
            if not df_final.empty:
                df_display = df_final.copy()
                mapa_colunas = {
                    'Descrição': ('Dados', 'Banco / Descrição'), 
                    'Chave Primaria': ('Dados', 'Conta Reduzida'),
                    'Saldo_Contabil_CC': ('Conta Corrente', 'Contábil'), 
                    'Saldo_Banco_CC': ('Conta Corrente', 'Banco'), 
                    'Diferenca_Saldo_CC': ('Conta Corrente', 'Diferença'),
                    'Saldo_Contabil_Aplic': ('Aplicação', 'Contábil'), 
                    'Saldo_Banco_Aplic': ('Aplicação', 'Banco'), 
                    'Diferenca_Saldo_Aplic': ('Aplicação', 'Diferença'),
                    'Rendimento_Contabil': ('Rendimentos', 'Contábil'), 
                    'Rendimento_Banco': ('Rendimentos', 'Banco'), 
                    'Diferenca_Rendimento': ('Rendimentos', 'Diferença')
                }
                
                cols_existentes = [c for c in df_display.columns if c in mapa_colunas]
                df_display = df_display[cols_existentes]
                df_display.columns = pd.MultiIndex.from_tuples([mapa_colunas[c] for c in df_display.columns])
                
                numeric_cols = df_display.select_dtypes(include=['float', 'int']).columns
                df_formatado = df_display.copy()
                for col in numeric_cols: df_formatado[col] = df_formatado[col].apply(formatar_moeda_br)

                st.success("Processamento concluído.")
                
                tab1, tab2, tab3, tab4 = st.tabs(["📊 Visão Geral", "🚨 Apenas Divergências", "📝 Log de Leitura", "🔄 Mapa De-Para"])
                
                with tab1:
                    st.dataframe(df_formatado, use_container_width=True, height=500)
                    col_dl1, col_dl2 = st.columns(2)
                    with col_dl1:
                        st.download_button("📥 Baixar Excel Formatado", to_excel_styled(df_display), "conciliacao_completa.xlsx", type='primary', use_container_width=True)
                    with col_dl2:
                        st.download_button("📄 Baixar Relatório PDF", to_pdf(df_display), "relatorio_conciliacao.pdf", use_container_width=True)
                
                with tab2:
                    filtro = (df_final['Diferenca_Saldo_CC'].abs() > 0.01) | \
                             (df_final['Diferenca_Saldo_Aplic'].abs() > 0.01) | \
                             (df_final['Diferenca_Rendimento'].abs() > 0.01)
                    df_div = df_formatado[filtro]
                    if df_div.empty: st.info("Tudo certo! Nenhuma divergência encontrada.")
                    else: st.dataframe(df_div, use_container_width=True)
                
                with tab3:
                    st.dataframe(df_log, use_container_width=True)
                    
                with tab4:
                    if not df_depara.empty:
                        st.success(f"✅ De-Para processado! {len(df_depara)} mapeamentos realizados.")
                        st.dataframe(df_depara, use_container_width=True)
                        st.info("⚠️ Se você ver duas contas separadas na Visão Geral que deveriam estar juntas, verifique as colunas **Chave Antiga** e **Chave Nova** abaixo. Os números precisam ser exatamente iguais aos 7 últimos dígitos gerados na coluna **Conta Reduzida** da Visão Geral.")
                    else:
                        st.error("⚠️ De-Para retornou VAZIO. Verifique as mensagens de erro.")
            else:
                st.error("O processamento não retornou dados.")
